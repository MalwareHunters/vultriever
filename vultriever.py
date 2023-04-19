# coding: utf-8

import re
import os
import sys
import subprocess
import openpyxl
import json
import time
import requests
import urllib3
from threading import Thread
from domain_ip_resolvers import *

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logo = """        
                                                                            
    ██╗   ██╗██╗   ██╗██╗  ████████╗██████╗ ██╗███████╗██╗   ██╗███████╗██████╗ 
    ██║   ██║██║   ██║██║  ╚══██╔══╝██╔══██╗██║██╔════╝██║   ██║██╔════╝██╔══██╗
    ██║   ██║██║   ██║██║     ██║   ██████╔╝██║█████╗  ██║   ██║█████╗  ██████╔╝
    ╚██╗ ██╔╝██║   ██║██║     ██║   ██╔══██╗██║██╔══╝  ╚██╗ ██╔╝██╔══╝  ██╔══██╗
     ╚████╔╝ ╚██████╔╝███████╗██║   ██║  ██║██║███████╗ ╚████╔╝ ███████╗██║  ██║
      ╚═══╝   ╚═════╝ ╚══════╝╚═╝   ╚═╝  ╚═╝╚═╝╚══════╝  ╚═══╝  ╚══════╝╚═╝  ╚═╝

                   IN MEMORY OF VINCENT. DEFENDER, HUNTER & FRIEND                             

⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠿⠿⠿⠿⠿⠿⠿⠿⠿⠿⠿⢿⣿⠿⠿⠿⠿⠿⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠛⢁⣤⣶⣶⣶⣶⣿⣿⣿⣷⣶⣶⣶⣶⣶⣶⣶⣶⣶⣶⣶⣶⣶⣶⣶⣶⣶⣤⣉⠻⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠏⢀⢴⣾⣿⠏⣽⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⠻⣷⣄⠙⢿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠋⡀⣡⣿⠿⠃⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣌⢈⠻⣷⣄⠙⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠟⠁⢂⠜⡩⠴⠀⢺⣿⣿⣿⣿⣿⣿⣿⣿⣿⣏⢻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣛⢿⣿⣿⣿⣿⣿⣦⠡⠙⢿⣿⣦⡈⠻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠟⠁⣰⣿⣿⠏⢠⡡⢠⣿⣿⣿⡿⠛⠉⢉⠉⠛⣿⣿⣾⣿⣿⣿⣿⣿⣿⣿⣿⡇⠀⡀⠀⠉⢙⠿⣿⣿⣿⣧⠡⠁⠹⣿⣷⣄⡈⠛⠿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⠿⠟⢃⣠⣾⡿⢋⡜⠀⣸⢃⣾⣿⣿⠏⠀⠀⠐⠀⠀⠄⠘⢁⣾⣿⣿⣿⣿⣿⣿⣿⣿⡇⢀⠀⠀⠂⠈⠐⡙⢿⣿⣿⡄⣧⠀⠈⠻⣿⣿⡂⠀⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⡦⠉⣴⣶⣿⣿⠿⣋⣴⣿⠀⢰⠃⣸⣿⣿⣷⡄⢀⠀⠀⠀⠀⠀⣸⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⡈⢀⣀⠀⠀⠤⣶⣾⣿⣷⡀⡈⠁⠀⣷⣌⡻⣿⣦⣁⠀⠀⠿⣿⣿⣿⣿⣿
⣿⣿⣿⡿⠃⢠⣿⣿⡿⣿⣿⣿⣿⡇⠀⡟⠀⣿⣿⣿⣿⣿⣷⣿⣿⡟⢁⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣌⢻⣿⣿⣿⣿⣿⣿⣿⣷⡄⠀⠀⣿⡟⢿⣷⣝⠿⣷⡀⠈⠙⠿⣿⣿⣿
⣿⣿⣿⠟⠠⣿⡿⣋⣴⢟⣿⡟⣹⣿⠀⡇⢰⣿⣿⣿⣿⣿⣿⣿⠟⣰⣿⣿⣿⣿⡿⠿⠛⠛⠛⠛⠻⢿⣿⣿⣿⣿⣦⠹⣿⣿⣿⣿⣿⣿⣿⣧⠀⠀⣿⣿⣦⡝⢌⢳⣌⠃⠀⠓⢶⣿⣿⣿
⣿⣿⢏⠀⢠⣿⡿⣫⣴⡿⢋⣴⣿⣿⠀⡇⢸⣿⣿⣿⣿⡿⠿⣋⣼⣿⣿⣿⡟⠁⠀⠀⠀⠀⠀⠀⠀⠀⠈⢻⣿⣿⣿⣧⣘⠿⣿⣿⣿⣿⣿⡿⠀⠀⡇⣿⣿⡝⣿⣦⡙⢷⠀⡈⢿⣿⣿⣿
⣿⣿⣁⡀⠈⢫⣾⡿⢋⣴⣾⣿⡿⠛⠀⠇⢸⣿⡟⣋⣭⣴⣾⣿⣿⣿⣿⣿⣇⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⢸⣿⣿⣿⣿⣿⣷⣬⣍⡛⢿⣿⡇⠀⠀⡇⣿⣯⠹⣌⠻⣿⣬⠀⢻⣿⣿⣿⣿
⣿⣿⣿⡆⠀⠘⣩⣾⡿⢋⣼⡿⢁⣦⠈⠀⢸⣿⠰⣿⡈⣿⣿⣿⣿⣿⣿⣿⣿⠀⠀⠀⠀⠀⠀⠀⠀⠀⠀⣼⣿⣿⣿⣿⣿⣿⣿⡿⢿⡜⣿⠁⠀⣰⢰⣿⣿⣑⠹⣧⡘⡿⠀⠈⣿⣿⣿⣿
⣿⣿⣿⣷⠀⠸⠛⣉⡴⢋⢟⠁⣾⣿⠀⠀⢸⣿⣧⣿⠁⠙⣿⣿⣿⣿⣿⣿⣿⣷⣤⣄⡀⠀⠀⠀⢀⣤⣾⣿⣿⣿⣿⣿⣿⡿⠋⠀⡹⡇⣿⠀⠀⠃⣿⣟⠻⣝⠳⣠⠱⠆⠀⠀⣿⣿⣿⣿
⣿⣿⣿⣿⡄⠀⢠⣿⠔⡡⣢⣾⣿⣿⠃⠆⠘⡿⢿⣿⠀⠀⠘⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⠁⠀⣙⣻⣿⣿⣿⣿⣿⣿⣿⡿⠀⠀⢠⢠⣿⡏⠀⢰⠀⣿⣿⡀⢌⠻⡟⠁⣀⢀⣠⣿⣿⣿⣿
⣿⣿⣿⣿⣷⣦⠀⠠⠾⢿⣿⠿⠟⡛⢠⠀⢰⣣⢛⣿⡼⡄⠀⠈⠻⢿⣿⣿⣿⣟⠿⠟⢃⡀⢀⠙⠯⠻⢿⣿⣿⣟⣛⠋⢠⠀⢀⢇⣿⣿⠁⠀⢸⢦⠹⣷⡙⠧⠑⠀⣄⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣶⡄⠀⡀⣡⠦⠀⡠⠁⠀⠸⡟⢸⣿⣇⢸⡄⠈⢶⣤⣤⠀⣤⣤⣶⣶⣿⡇⢸⣿⣶⣶⣤⡄⠀⠀⣠⣶⣿⠁⢸⣿⣿⡟⠀⠀⠘⡟⢢⢸⡇⢠⡀⣼⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⡆⠡⡀⠀⠟⠁⠀⠀⠀⠀⡾⣿⣿⣆⠳⠄⢠⣿⡆⠀⢻⣿⣿⣿⣿⡇⣿⣿⣿⣿⡿⠀⡄⢰⣿⣿⠄⢀⣿⣿⢯⣇⠀⠀⠀⢿⣆⠘⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣶⣦⡤⠖⠀⣆⣄⡀⡸⢃⣿⣿⣿⣆⢰⡀⢻⣷⠀⢸⣿⣿⣿⣿⣧⣿⣿⣿⣿⡇⠐⢀⣿⣿⠏⠀⣾⠻⢿⠀⢻⣦⣷⡀⠀⠙⢀⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⢰⡇⢸⣿⣏⢁⢡⡾⠙⣿⣿⣿⡘⠷⡈⢻⣦⠀⢿⣿⣿⣿⣿⣿⣿⣿⣿⠃⢠⣾⣿⠋⠀⣼⣿⠎⠈⢀⢸⣿⢸⡇⠹⣆⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⢸⡇⢸⣿⣿⣸⢸⠇⢐⣿⣿⣿⣷⡄⢿⣄⠻⣇⠈⢿⣿⣿⣿⣿⣿⣿⠋⢠⡿⠛⠁⢀⣼⣿⣾⣆⣴⠘⣿⣿⣾⡿⢠⢻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠘⡃⢸⣿⣿⣿⡜⢠⠟⠏⢼⣿⣿⣿⣎⢻⣦⠙⢧⡀⠙⠻⠿⠿⠟⢁⡴⠋⠀⠀⣠⣿⠿⣿⡇⣿⣿⣤⣾⣿⡿⠁⢸⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣆⠃⢈⢿⣿⣿⡇⠈⡤⠀⠸⢹⣿⣿⣿⣧⣙⠳⣄⠉⠓⠶⣶⣶⠿⠋⠀⣀⣠⣪⣾⡟⡄⢠⠀⢩⣿⣿⣿⣿⢃⣶⠈⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡘⡎⠈⣿⣿⣿⣾⠀⡆⣴⠀⣼⣿⣿⣿⣿⣷⣬⡃⢄⠐⠒⠒⠀⡠⢞⣫⣽⣿⣿⣿⠁⠀⠀⢸⣿⣿⣿⡟⢸⢷⢰⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣇⢰⡀⠸⣿⣿⣿⣤⡀⠏⠀⠁⢹⣿⣿⣿⡿⣿⣿⣷⣾⣭⣭⣭⣶⣿⣿⣿⣟⣿⣥⠀⣄⠙⣿⣿⣿⡿⢠⡟⠈⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣄⠓⣄⢹⣿⣿⣿⣇⣠⣁⣿⠈⡃⠹⣿⣷⢸⠏⡀⠛⣿⣿⠿⡿⣿⣿⠏⠀⣺⡟⡀⢹⣾⣿⣿⡿⢡⡿⡐⢠⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣦⠘⢯⣿⣿⣿⣿⣿⠃⠋⡀⣼⣿⣿⠙⠈⣰⡕⠀⠻⢁⠀⠀⣿⢃⣔⣰⠏⣰⣿⣿⣿⣿⣿⢡⣿⠐⣱⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣦⠻⣿⣿⣿⣿⣿⣿⣷⣿⣿⣷⣷⣾⣿⣾⠀⠰⣿⠀⠡⣸⣿⡿⣁⣾⣿⣿⣽⣿⣿⣷⠟⣡⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣬⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠋⣴⣧⣿⣧⣶⣿⣋⣴⣿⣿⣿⣿⣿⣿⣿⣋⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿

OFFICIAL SITE:          www.malwarehunters.org
TELEGRAM CHANNEL:       t.me/malwarehunters
TELEGRAM CONTACT:       t.me/malware_hunters
TWITTER:                twitter.com/@_malwarehunters
EMAIL:                  threat@malwarehunters.org

⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
"""

help_message = """
VULTRIEVER uses the standard command syntax for Nmap scanner (see below, need to install) and several unique parameters:

    -CVE    Activate a vulnerability checker that defines a CVE identifier, 
            score and description of a well-known vulnerabilities for each detected service
    
    -HARV   Activate the use of Harvester (need to install) utility to detect subdomains of target 
            resources and their subsequent analysis in accordance with the specified parameters

    -RES    Activate the module for collecting additional information about the infrastructure: 
            countries, cities, hosting services, autonomous systems number of target servers
            and registrars of target domains
    
    -TITLE  Activate the target resources titles collection module    

VULTRIEVER usage in TERMINAL: 

    # sudo python vultriever.py <Nmap arguments string>
    
    # sudo python vultriever.py x.x.x.x y.y.y.y z.z.z.z/24 -sS -sU --top-ports 100 -Pn -CVE -HARV -RES -TITLE
    # sudo python vultriever.py domain1.com domain2.com -sS -sU --top-ports 100 -Pn -CVE -HARV -RES -TITLE
    # sudo python vultriever.py x.x.x.x domain1.com z.z.z.z/24 -sS -sU --top-ports 100 -Pn -CVE -HARV -RES -TITLE
    # sudo python vultriever.py -iL targets.txt -sS -sU --top-ports 100 -Pn -CVE -HARV -RES -TITLE
    
    Input targets list format:
    
        _______________
        |x.x.x.x      |
        |y.y.y.y      |
        |             |
        |. . . .      |
        |             |
        |z.z.z.z/24   |
        |domain1.com  |
        |             |
        |. . . .      |
        |             |
        |domain100.com|
        ---------------

    The results of the analysis are stored in a structured form in Excel file

VULTRIEVER usage in PYTHON CODE:    

    vultriever(['<target>', ... , '<target>'], '<Nmap arguments string>')

    ----------------- PYTHON SCRIPT -----------------
    
    from vultriever import vultriever

    analysis_json = vultriever(['x.x.x.x', 'y.y.y.y', 'z.z.z.z/24'], '-sS -sU --top-ports 100 -Pn -CVE -HARV -RES -TITLE')
    analysis_json = vultriever(['domain1.com', 'domain2.com'], '-sS -sU --top-ports 100 -Pn -CVE -HARV -RES -TITLE')
    analysis_json = vultriever(['x.x.x.x', 'domain1.com', 'z.z.z.z/24'], '-sS -sU --top-ports 100 -Pn -CVE -HARV -RES -TITLE')
    
    -------------------------------------------------

    vultriever() function returns information in JSON format with next structure:

    {
        "ip_address" : "x.x.x.x",
        "country": "United States",
        "city": "San Francisco",
        "hosting": "Cloudflare, Inc.",
        "asn": "13335",
        
        "domain": "domain.com",
        "title": "Target Domain",
        "registrar": "Regional Network Information Center, JSC dba RU-CENTER",

        "ports" : [
            
                {
                    "number" : 26,
                    "status" : "open",
                    "protocol" : "smtp",
                    "service" : "Eximsmtpd4.95",
                    "vulnerabilities" : [
                        
                                {
                                    "cve" : "CVE-2022-37452",
                                    "score" : "9.8",
                                    "url" : "https://vulners.com/cve/CVE-2022-37452"
                                }
                        
                    ]
                    
                }
            
        ]
    }

    Incorrect request will result to 'error_message' in JSON response keys
    
    {
        
        "error_message" : <error message>
        
    }

⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿
"""

class Server(object):

    def __init__(self, ip_address):
        
        self.ip_address = ip_address
        self.ports = []
        self.country = None
        self.city = None
        self.hosting = None
        self.asn = None
        self.ports = []
        self.title = None

    def resolve(self):
        
        ip_info = ip_resolver(self.ip_address)
        self.country = ip_info.get('country')
        self.city = ip_info.get('city')
        self.hosting = ip_info.get('provider')
        self.asn = ip_info.get('asn') 

    def get_title(self):
        
        try:
            r = requests.get('http://{0}'.format(self.ip_address), verify=False, timeout=15)
            text = r.text
            self.title = text[text.find('<title>') + 7 : text.find('</title>')]
        except: pass
    
class Domain(object):

    def __init__(self, domain):
    
        if len(domain) > 4:
            if domain[0:4] == 'www.': self.domain = domain[4:]
            else: self.domain = domain
        else: self.domain = domain
        self.title = None
        self.registrar = None

    def resolve(self):
        
        self.registrar = domain_resolver(self.domain).get('registrar')

    def get_title(self):
        
        try:
            r = requests.get('https://{0}'.format(self.domain), verify=False, timeout=15)
            text = r.text
            self.title = text[text.find('<title>') + 7 : text.find('</title>')]
        except: pass

class Port(object):
    
    def __init__(self, number):
        
        self.number = number
        self.status = None
        self.protocol = None
        self.service = None
        self.vulnerabilities = []

    
class Vuln(object):
    
    def __init__(self, cve):
        
        self.cve = cve
        self.score = None
        self.url = None


class Target(object):
    
    def __init__(self, server = None, domain = None): 

        if server: self.server = server
        else: self.server = None

        if domain: self.domain = domain
        else: self.domain = None

class NMap_Collector(object):

    servers_pattern = re.compile('Nmap scan report for .+(?:\n[\S| |\t]+)+')
    server_pattern = re.compile('\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}')
    domain_pattern = re.compile('[A-Za-z0-9-]+\.[A-Za-z0-9-\.]+')
    ports_pattern = re.compile('\d+/[tcpud]+[\s| |\t]+[open|closed|filtered|unfiltered]+[\S| |\t]+(?:\n\|[\S| |\t]+)*')
    port_pattern = re.compile('\d+')
    CVE_pattern = re.compile('\|[ |\t]+CVE-\d+-\d+[\t]+\d.\d[\t]+\S+')
    illegal_mask_error_pattern = re.compile('Illegal netmask .+')
    failed_resolve_error_pattern = re.compile('Failed to resolve .+')
    
    def __init__(self, nmap_result):

        self.targets = []

        nmap_result = nmap_result.replace('\nPORT','PORT')
        servers_info = re.findall(self.servers_pattern, nmap_result)
        for server_info in servers_info:
        
            server = Server(re.findall(self.server_pattern, server_info)[0])

            domains = re.findall(self.domain_pattern, server_info)
            if domains: 
                if domains[0] != server.ip_address:
                    domain = Domain(domains[0])
                    if domain.domain in self.get_domains(): continue
                else:
                    domain = None
            else: 
                domain = None
            
            for port_info in re.findall(self.ports_pattern, server_info):
            
                port = Port(re.findall(self.port_pattern, port_info)[0])
                                
                port_status = list(filter(None, port_info.split('\n')[0].split(' ')))
                if len(port_status) >= 2: port.status = port_status[1]
                if len(port_status) >= 3: port.protocol = port_status[2]
                if len(port_status) >= 4: port.service = ' '.join(port_status[3:])
                
                CVEs_info = re.findall(self.CVE_pattern, port_info)
                
                if CVEs_info:

                    for CVE_info in CVEs_info:
                    
                        info = CVE_info.split('\t')  
                        vuln = Vuln(info[1])  
                        vuln.score = info[2]
                        vuln.url = info[3]

                        port.vulnerabilities.append(vuln)
    
                server.ports.append(port)                                            
                                                                        
            self.targets.append(Target(server=server, domain=domain))
                    
    def ips_resolve(self):
    
        for target in self.targets:
        
            if target.server: target.server.resolve()

    def domains_resolve(self):
    
        for target in self.targets:
        
            if target.domain: target.domain.resolve()

    def titles_collect(self):
        
        for target in self.targets:
        
            if target.domain: target.domain.get_title()
            else: target.server.get_title()    

    def get_domains(self):
        
        domains = []
        
        for target in self.targets:
        
            if target.domain: domains.append(target.domain.domain)
            
        return list(set(domains))
            
    def export_json(self):
        
        json_response = []
        for target in self.targets:
        
            json_full = {

                        "ip_address": target.server.ip_address,
                        "country": target.server.country,
                        "city": target.server.city,
                        "hosting": target.server.hosting,
                        "asn": target.server.asn,
                        "ports": [],

                        "domain": target.domain.domain if target.domain else '',
                        "title": target.domain.title if target.domain else '',
                        "registrar": target.domain.registrar if target.domain else ''
                    }

            for port in target.server.ports:
            
                json_port = {

                            "number": port.number,
                            "status": port.status,
                            "protocol": port.protocol,
                            "service": port.service,
                            "vulnerabilities": []
                    
                }    
                
                for vuln in port.vulnerabilities:
                
                    vuln_json = {
                        
                                "cve": vuln.cve,
                                "score": vuln.score,
                                "url": vuln.url
                        
                    }
                    
                    json_port['vulnerabilities'].append(vuln_json)
                    
                json_full['ports'].append(json_port)
                
            json_response.append(json_full)
            
        return json.dumps(json_response)
        
    def export_excel(self, excel_filename = None):

        if excel_filename:
        
            excel_book = openpyxl.load_workbook(excel_filename)
            excel_sheet = excel_book.active
            row_num = excel_sheet.max_row + 1
        
        else:

            excel_book = openpyxl.Workbook()
            del excel_book['Sheet']
            excel_sheet = excel_book.create_sheet('CVE Info')
            excel_sheet.cell(row=1,column=1,value='Domain')
            excel_sheet.cell(row=1,column=2,value='Title')
            excel_sheet.cell(row=1,column=3,value='Registrar')
            excel_sheet.cell(row=1,column=4,value='IP')
            excel_sheet.cell(row=1,column=5,value='Country')
            excel_sheet.cell(row=1,column=6,value='City')
            excel_sheet.cell(row=1,column=7,value='Hosting')
            excel_sheet.cell(row=1,column=8,value='ASN')
            excel_sheet.cell(row=1,column=9,value='Port')
            excel_sheet.cell(row=1,column=10,value='Status')
            excel_sheet.cell(row=1,column=11,value='Protocol')
            excel_sheet.cell(row=1,column=12,value='Service')
            excel_sheet.cell(row=1,column=13,value='CVE')
            excel_sheet.cell(row=1,column=14,value='Score')
            excel_sheet.cell(row=1,column=15,value='URL')
        
            row_num = 2
        
        for target in self.targets:
        
            for port in target.server.ports:
        
                if port.vulnerabilities:
                
                    for vuln in port.vulnerabilities:
                
                        excel_sheet.cell(row=row_num,column=1,value=target.domain.domain if target.domain else '')
                        excel_sheet.cell(row=row_num,column=2,value=target.domain.title if target.domain else target.server.title)
                        excel_sheet.cell(row=row_num,column=3,value=target.domain.registrar if target.domain else '')
                        excel_sheet.cell(row=row_num,column=4,value=target.server.ip_address)
                        excel_sheet.cell(row=row_num,column=5,value=target.server.country)
                        excel_sheet.cell(row=row_num,column=6,value=target.server.city)
                        excel_sheet.cell(row=row_num,column=7,value=target.server.hosting)
                        excel_sheet.cell(row=row_num,column=8,value=target.server.asn)
                        excel_sheet.cell(row=row_num,column=9,value=port.number)
                        excel_sheet.cell(row=row_num,column=10,value=port.status)
                        excel_sheet.cell(row=row_num,column=11,value=port.protocol)
                        excel_sheet.cell(row=row_num,column=12,value=port.service)
                        excel_sheet.cell(row=row_num,column=13,value=vuln.cve)
                        excel_sheet.cell(row=row_num,column=14,value=vuln.score)
                        excel_sheet.cell(row=row_num,column=15,value=vuln.url)
                        row_num += 1
                
                else:
                                    
                    excel_sheet.cell(row=row_num,column=1,value=target.domain.domain if target.domain else '')
                    excel_sheet.cell(row=row_num,column=2,value=target.domain.title if target.domain else target.server.title)
                    excel_sheet.cell(row=row_num,column=3,value=target.domain.registrar if target.domain else '')
                    excel_sheet.cell(row=row_num,column=4,value=target.server.ip_address)
                    excel_sheet.cell(row=row_num,column=5,value=target.server.country)
                    excel_sheet.cell(row=row_num,column=6,value=target.server.city)
                    excel_sheet.cell(row=row_num,column=7,value=target.server.hosting)
                    excel_sheet.cell(row=row_num,column=8,value=target.server.asn)
                    excel_sheet.cell(row=row_num,column=9,value=port.number)
                    excel_sheet.cell(row=row_num,column=10,value=port.status)
                    excel_sheet.cell(row=row_num,column=11,value=port.protocol)
                    excel_sheet.cell(row=row_num,column=12,value=port.service)
                    row_num += 1
    
        print('\nResult saved to Vultriever.xlsx')
        excel_book.save('Vultriever.xlsx')

class Harvester_Collector(object):
    
    def __init__(self, harvester_result):

        self.targets = []
        
        json_result = json.loads(harvester_result)
        if json_result.get('hosts'):
            for domain in json_result.get('hosts'):
                self.targets.append(Target(server=None,domain=Domain(domain.split(':')[0])))
            
    def get_domains(self):
        
        domains = []
        for target in self.targets:
            domains.append(target.domain.domain)

        return list(set(domains))


def vultriever(targets=None, options_string=None):

    def command_line_check(targets, options_string, sysargs=None):

        TERMINAL_MODE = True
        CVE_CHECKER_USAGE = False
        HARVERSTER_USAGE = False
        RESOLVER_USAGE = False
        TITLE_CRAWLER_USAGE = False
        
        if targets or options_string: 

            command_line = " ".join([" ".join(targets), options_string])
            TERMINAL_MODE = False

        else:
        
            TERMINAL_MODE = True
            
            os.system('clear')
            print(logo)

            
            if len(sys.argv) > 1:
            
                if sys.argv[1] == '--help':
                
                    print(help_message)
                    os.system('nmap --help')
                    sys.exit()
                
                else: command_line = " ".join(sys.argv[1:])
                
            else: 
            
                print('You must specify the parameters!')
                sys.exit()
        
        if '-HARV' in command_line: 
            HARVERSTER_USAGE = True
            command_line = command_line.replace('-HARV','').strip()
            if TERMINAL_MODE: print('[+] HARVESTER activated')

        if '-RES' in command_line: 
            RESOLVER_USAGE = True
            command_line = command_line.replace('-RES','').strip()
            if TERMINAL_MODE: print('[+] IP & Domains RESOLVER activated')

        if '-TITLE' in command_line: 
            TITLE_CRAWLER_USAGE = True
            command_line = command_line.replace('-TITLE','').strip()
            if TERMINAL_MODE: print('[+] TITLE CRAWLER activated')

        if '-CVE' in command_line: 

            command_line = command_line.replace('-CVE','').strip()

            if '--script=vulners.nse' not in command_line and '--script vulners' not in command_line:
                command_line += " --script vulners"
                
            if '-sV' not in command_line:
                command_line += " -sV"

            CVE_CHECKER_USAGE = True                        
            if TERMINAL_MODE: print('[+] CVE CHECKER activated')

        if '-n' not in command_line: command_line += " -n"

        return command_line.strip(), TERMINAL_MODE, CVE_CHECKER_USAGE, HARVERSTER_USAGE, RESOLVER_USAGE, TITLE_CRAWLER_USAGE

    try:
        
        command_line, TERMINAL_MODE, CVE_CHECKER_USAGE, HARVERSTER_USAGE, RESOLVER_USAGE, TITLE_CRAWLER_USAGE = command_line_check(targets, options_string, sys.argv)
        
        if command_line:

            try:

                os.chdir(os.getcwd())
                with open('nmap_log.txt', 'w') as nmap_log:
                    
                    nmap_command = "nmap " + command_line
                    if TERMINAL_MODE: print('\nStarting NMap scan: ' + nmap_command)
                    nmap_process = subprocess.call(nmap_command, shell=True, stdout=nmap_log)
                    if TERMINAL_MODE: print('\tNMap scan completed ...')

                with open('nmap_log.txt', 'r') as nmap_log:

                    nmap_collector = NMap_Collector(nmap_log.read())
                    if TERMINAL_MODE: print('\tNMap scan result analyze completed ...')
                    
                #HARVESTER
                if HARVERSTER_USAGE:
                
                    domains = nmap_collector.get_domains()
                    if TERMINAL_MODE: print('\nStarting theHarvester searching for {0} domain(s) ...'.format(len(domains)))
                    with open('targets_from_harvester.txt', 'w') as harvester_targets:
    
                        with open('harvester_log.txt', 'w') as harvester_log:
    
                            for domain in domains:
    
                                harvester_command = "theHarvester -d {} -b all -f harvester_log".format(domain)
                                if TERMINAL_MODE: print('\tUse theHarvester with ' + domain + '...')
                                harvester_process = subprocess.call(harvester_command, shell=True, stdout=harvester_log)
    
                                with open('harvester_log.json','r') as harvester_result:
    
                                    harvester_collector = Harvester_Collector(harvester_result.read())
                                    harvester_domains = harvester_collector.get_domains()
                                    if TERMINAL_MODE: print('\t\t{0} subdomain(s) detected ...'.format(len(harvester_domains)))
                                    for domain in harvester_domains:
                                        harvester_targets.write(domain+'\n')

                            if TERMINAL_MODE: print('\ttheHarvester searching completed ...')
                                    
                    with open('nmap_log.txt', 'a') as nmap_log:
            
                        nmap_command = "nmap -iL targets_from_harvester.txt " + command_line
                        if TERMINAL_MODE: print('\nStarting NMap scan for all domains: ' + nmap_command)
                        nmap_process = subprocess.call(nmap_command, shell=True, stdout=nmap_log)
                        if TERMINAL_MODE: print('\tNMap scan completed ...')

                    with open('nmap_log.txt', 'r') as nmap_log:
                    
                        nmap_collector = NMap_Collector(nmap_log.read())
                        if TERMINAL_MODE: print('\tNMap scan result analyze completed ...')
                        
                #RESOLVER
                if RESOLVER_USAGE: 
                
                    if TERMINAL_MODE: print('\nStarting IPs resolving ...')
                    nmap_collector.ips_resolve()

                    if TERMINAL_MODE: print('\tIPs resolving completed...')
                    
                    if TERMINAL_MODE: print('\nStarting domains resolving ...')
                    nmap_collector.domains_resolve()
                    if TERMINAL_MODE: print('\tDomains resolving completed...')
                    
                if TITLE_CRAWLER_USAGE:
                    if TERMINAL_MODE: print('\nStarting titles crawling ...')
                    nmap_collector.titles_collect()
                    if TERMINAL_MODE: print('\tTitles crawling completed...')
                    
            except Exception as error: 
                
                if not TERMINAL_MODE: return json.dumps({'error_message': str(error)})
                
                else:
                
                    print(error)
                    sys.exit()

            ### RESPONSE
            
            if not TERMINAL_MODE:
            
                ### JSON
            
                try:

                    return nmap_collector.export_json()
                    
                except Exception as error:
                
                    return json.dumps({'error_message':str(error)})
            
            else:

                ### EXCEL
                nmap_collector.export_excel()

        
    except Exception as error: print(error)

if __name__ == '__main__':

    vultriever()